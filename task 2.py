from openpyxl import load_workbook

# Load workbook and select the sheet
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

count = 0

for row in range(2, max_row + 1):
    priority = ws.cell(row=row, column=8).value       # Column H = 8
    delivery_date = ws.cell(row=row, column=10).value # Column J = 10

    if priority == 'High' and hasattr(delivery_date, 'year') and delivery_date.year == 2015:
        count += 1

print(count)
